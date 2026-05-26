"""
Generador de reportes Excel con formato profesional
Compatible con Microsoft 365 / Google Sheets
"""
import datetime
import io
from pathlib import Path
from openpyxl import Workbook
<<<<<<< HEAD
from openpyxl.cell import WriteOnlyCell
=======
>>>>>>> 7761f488b2aa6200974f069ea5072699c6dbd1e5
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from sqlalchemy.orm import Session
<<<<<<< HEAD
from sqlalchemy import case, func
=======
>>>>>>> 7761f488b2aa6200974f069ea5072699c6dbd1e5
from app.database import Cobro, Prestamo, Cliente, Cuota, Zona


# Paleta de colores
COLOR_VERDE_OSCURO = "0F3D28"
COLOR_VERDE = "1A6B4A"
COLOR_VERDE_CLARO = "E8F5EF"
COLOR_GRIS = "F5F5F5"
COLOR_ROJO = "C0392B"
COLOR_AMARILLO = "F9E79F"
COLOR_BLANCO = "FFFFFF"
COLOR_TEXTO_OSCURO = "1A1A1A"


def estilo_header(ws, fila: int, cols: list, labels: list, color_fondo=COLOR_VERDE_OSCURO):
    fill = PatternFill(start_color=color_fondo, end_color=color_fondo, fill_type="solid")
    font = Font(color=COLOR_BLANCO, bold=True, size=10, name="Calibri")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borde = Border(
        bottom=Side(style="medium", color=COLOR_VERDE),
        top=Side(style="thin", color=color_fondo),
    )
    for i, (col, label) in enumerate(zip(cols, labels), start=1):
        cell = ws.cell(row=fila, column=i, value=label)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = borde


def estilo_fila(ws, fila: int, valores: list, par: bool = True):
    bg = COLOR_GRIS if par else COLOR_BLANCO
    fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    font = Font(size=9, name="Calibri", color=COLOR_TEXTO_OSCURO)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    for i, val in enumerate(valores, start=1):
        cell = ws.cell(row=fila, column=i, value=val)
        cell.fill = fill
        cell.font = font
        cell.alignment = align_center if isinstance(val, (int, float)) else align_left
        cell.border = Border(bottom=Side(style="hair", color="DDDDDD"))


def encabezado_reporte(ws, titulo: str, subtitulo: str, empresa: str):
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 16

    # Logo / título
    fill_top = PatternFill(start_color=COLOR_VERDE_OSCURO, end_color=COLOR_VERDE_OSCURO, fill_type="solid")
    ws["A1"] = f"◆ {empresa}"
    ws["A1"].font = Font(color=COLOR_BLANCO, bold=True, size=16, name="Calibri")
    ws["A1"].fill = fill_top
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws["A2"] = titulo
    ws["A2"].font = Font(color=COLOR_VERDE, bold=True, size=12, name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    ws["A3"] = f"Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}  |  {subtitulo}"
    ws["A3"].font = Font(color="888888", italic=True, size=9, name="Calibri")


def reporte_cobros_diarios(db: Session, empresa_id: int = None, zona_id: int = None, fecha: datetime.date = None) -> bytes:
    if fecha is None:
        fecha = datetime.date.today()

    wb = Workbook()
    ws = wb.active
    ws.title = "Cobros del Día"
    ws.sheet_view.showGridLines = False

    # Config
<<<<<<< HEAD
    empresa = "CreditosPro"
    try:
        from app.database import ConfiguracionApp
        _cfg2 = db.query(ConfiguracionApp).filter(ConfiguracionApp.empresa_id==empresa_id).first() if empresa_id else None
        if _cfg2 and _cfg2.empresa_nombre: empresa = _cfg2.empresa_nombre
    except Exception:
        empresa = "CreditosPro"
=======
    from app.database import ConfiguracionApp
    config = db.query(ConfiguracionApp).first()
    empresa = config.empresa_nombre if config else "CreditosPro"
>>>>>>> 7761f488b2aa6200974f069ea5072699c6dbd1e5

    encabezado_reporte(ws, "REGISTRO DE COBROS DIARIOS", f"Fecha: {fecha.strftime('%d/%m/%Y')}", empresa)

    # Merge header
    ws.merge_cells("A1:J1")
    ws.merge_cells("A2:J2")
    ws.merge_cells("A3:J3")

<<<<<<< HEAD
    # Query - siempre filtrar por empresa_id cuando se proporciona
=======
    # Query
>>>>>>> 7761f488b2aa6200974f069ea5072699c6dbd1e5
    query = db.query(Cobro).filter(Cobro.fecha == fecha)
    if empresa_id:
        query = query.filter(Cobro.empresa_id == empresa_id)
    if zona_id:
        query = query.filter(Cobro.zona_id == zona_id)
    cobros = query.all()

<<<<<<< HEAD
    # Precargar datos relacionados en pocos queries (anti-N+1)
    cobro_ids = [c.id for c in cobros]
    clientes_ids = list({c.cliente_id for c in cobros})
    zona_ids = list({c.zona_id for c in cobros})
    cuota_ids = list({c.cuota_id for c in cobros if c.cuota_id})

    clientes_map = {}
    if clientes_ids:
        for cl in db.query(Cliente).filter(Cliente.id.in_(clientes_ids)).all():
            clientes_map[cl.id] = cl

    zonas_map = {}
    if zona_ids:
        for z in db.query(Zona).filter(Zona.id.in_(zona_ids)).all():
            zonas_map[z.id] = z

    cuotas_map = {}
    if cuota_ids:
        for cu in db.query(Cuota).filter(Cuota.id.in_(cuota_ids)).all():
            cuotas_map[cu.id] = cu

=======
>>>>>>> 7761f488b2aa6200974f069ea5072699c6dbd1e5
    fila_header = 5
    labels = ["#", "Cliente", "Cédula", "Zona", "Cuota N°", "Valor Cobrado", "Método Pago", "Cobrador", "Hora", "Observaciones"]
    estilo_header(ws, fila_header, list(range(1, 11)), labels)
    ws.row_dimensions[fila_header].height = 22

    total_cobrado = 0
    for idx, c in enumerate(cobros, start=1):
<<<<<<< HEAD
        cliente = clientes_map.get(c.cliente_id)
        zona = zonas_map.get(c.zona_id)
        cuota = cuotas_map.get(c.cuota_id)
=======
        cliente = db.query(Cliente).filter(Cliente.id == c.cliente_id).first()
        zona = db.query(Zona).filter(Zona.id == c.zona_id).first()
        cuota = db.query(Cuota).filter(Cuota.id == c.cuota_id).first()
>>>>>>> 7761f488b2aa6200974f069ea5072699c6dbd1e5

        fila = fila_header + idx
        vals = [
            idx,
            cliente.nombre if cliente else "—",
            cliente.cedula if cliente else "—",
            zona.nombre if zona else "—",
            cuota.numero if cuota else "—",
            c.valor_cobrado,
            c.metodo_pago,
            c.cobrador,
            c.hora.strftime("%H:%M") if c.hora else "—",
            c.observaciones or "",
        ]
        estilo_fila(ws, fila, vals, par=(idx % 2 == 0))
        total_cobrado += c.valor_cobrado

        # Colorear por monto
        cell_val = ws.cell(row=fila, column=6)
        cell_val.number_format = '#,##0'

    # Total
    fila_total = fila_header + len(cobros) + 1
    ws.cell(row=fila_total, column=5, value="TOTAL:").font = Font(bold=True, size=10, name="Calibri")
    cell_total = ws.cell(row=fila_total, column=6, value=total_cobrado)
    cell_total.font = Font(bold=True, size=11, color=COLOR_VERDE, name="Calibri")
    cell_total.number_format = '#,##0'

    # Anchos de columna
    anchos = [5, 28, 15, 15, 10, 16, 14, 22, 10, 30]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def reporte_cartera(db: Session, empresa_id: int = None) -> bytes:
<<<<<<< HEAD
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Cartera Activa")

    empresa = "CreditosPro"
    try:
        from app.database import ConfiguracionApp
        _cfg2 = db.query(ConfiguracionApp).filter(ConfiguracionApp.empresa_id==empresa_id).first() if empresa_id else None
        if _cfg2 and _cfg2.empresa_nombre: empresa = _cfg2.empresa_nombre
    except Exception:
        empresa = "CreditosPro"

    ws.append([empresa])
    ws.append(["ESTADO DE CARTERA"])
    ws.append([f"Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')} | Saldos vigentes por cliente"])
    ws.append([])

    estados_activos = ["Activo", "activo", "Atrasado", "atrasado", "Mora", "mora"]
    q = (
        db.query(
            Prestamo.id.label("prestamo_id"),
            Prestamo.capital,
            Prestamo.total_pagar,
            Prestamo.num_cuotas,
            Prestamo.estado,
            Cliente.nombre.label("cliente_nombre"),
            Cliente.cedula.label("cliente_cedula"),
            Zona.nombre.label("zona_nombre"),
            func.coalesce(func.sum(Cuota.valor_pagado), 0).label("pagado"),
            func.coalesce(func.sum(case((Cuota.estado == "Pagada", 1), else_=0)), 0).label("al_dia"),
            func.coalesce(func.sum(case((Cuota.estado == "Vencida", 1), else_=0)), 0).label("vencidas"),
            func.min(case((Cuota.estado == "Pendiente", Cuota.fecha_vencimiento), else_=None)).label("prox"),
        )
        .join(Cliente, Cliente.id == Prestamo.cliente_id)
        .outerjoin(Zona, Zona.id == Prestamo.zona_id)
        .outerjoin(Cuota, Cuota.prestamo_id == Prestamo.id)
        .filter(Prestamo.estado.in_(estados_activos))
    )
    if empresa_id:
        q = q.filter(Prestamo.empresa_id == empresa_id)
    prestamos = (
        q.group_by(
            Prestamo.id,
            Prestamo.capital,
            Prestamo.total_pagar,
            Prestamo.num_cuotas,
            Prestamo.estado,
            Cliente.nombre,
            Cliente.cedula,
            Zona.nombre,
        )
        .order_by(Cliente.nombre.asc(), Prestamo.id.asc())
        .all()
    )

    labels = ["Cliente", "Cédula", "Zona", "Capital", "Total", "Pagado", "Saldo", "Cuotas", "Al día", "Vencidas", "Próx. vencimiento", "Estado"]
    header_fill = PatternFill(start_color=COLOR_VERDE_OSCURO, end_color=COLOR_VERDE_OSCURO, fill_type="solid")
    header_font = Font(color=COLOR_BLANCO, bold=True, size=10, name="Calibri")
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_row = []
    for label in labels:
        cell = WriteOnlyCell(ws, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        header_row.append(cell)
    ws.append(header_row)

    total_capital = total_saldo = 0

    for idx, p in enumerate(prestamos, start=1):
        total_pagar = float(p.total_pagar or p.capital or 0)
        capital = float(p.capital or 0)
        pagado = float(p.pagado or 0)
        saldo = max(0, total_pagar - pagado)
        al_dia = int(p.al_dia or 0)
        vencidas = int(p.vencidas or 0)

        vals = [
            p.cliente_nombre,
            p.cliente_cedula,
            p.zona_nombre or "—",
            capital,
            total_pagar,
=======
    wb = Workbook()
    ws = wb.active
    ws.title = "Cartera Activa"
    ws.sheet_view.showGridLines = False

    from app.database import ConfiguracionApp
    config = db.query(ConfiguracionApp).first()
    empresa = config.empresa_nombre if config else "CreditosPro"

    encabezado_reporte(ws, "ESTADO DE CARTERA", "Saldos vigentes por cliente", empresa)
    ws.merge_cells("A1:L1")
    ws.merge_cells("A2:L2")
    ws.merge_cells("A3:L3")

    q = db.query(Prestamo).filter(Prestamo.estado.in_(["Activo", "Atrasado", "Mora"]))
    if empresa_id:
        q = q.filter(Prestamo.empresa_id == empresa_id)
    prestamos = q.all()

    labels = ["Cliente", "Cédula", "Zona", "Capital", "Total", "Pagado", "Saldo", "Cuotas", "Al día", "Vencidas", "Próx. vencimiento", "Estado"]
    estilo_header(ws, 5, list(range(1, 13)), labels)

    hoy = datetime.date.today()
    total_capital = total_saldo = 0

    for idx, p in enumerate(prestamos, start=1):
        cliente = p.cliente
        zona = db.query(Zona).filter(Zona.id == p.zona_id).first()
        pagado = sum(c.valor_pagado for c in p.cuotas)
        saldo = max(0, p.total_pagar - pagado)
        al_dia = sum(1 for c in p.cuotas if c.estado == "Pagada")
        vencidas = sum(1 for c in p.cuotas if c.estado == "Vencida")
        prox = min(
            (c.fecha_vencimiento for c in p.cuotas if c.estado == "Pendiente"),
            default=None
        )

        fila = 5 + idx
        vals = [
            cliente.nombre,
            cliente.cedula,
            zona.nombre if zona else "—",
            p.capital,
            p.total_pagar,
>>>>>>> 7761f488b2aa6200974f069ea5072699c6dbd1e5
            pagado,
            saldo,
            f"{al_dia + vencidas}/{p.num_cuotas}",
            al_dia,
            vencidas,
<<<<<<< HEAD
            p.prox.strftime("%d/%m/%Y") if p.prox else "—",
            p.estado,
        ]
        ws.append(vals)

        total_capital += capital
        total_saldo += saldo

    ws.append([])
    ws.append(["", "", "TOTALES:", total_capital, "", "", total_saldo])
=======
            prox.strftime("%d/%m/%Y") if prox else "—",
            p.estado,
        ]
        estilo_fila(ws, fila, vals, par=(idx % 2 == 0))

        for col in [4, 5, 6, 7]:
            ws.cell(row=fila, column=col).number_format = '#,##0'

        # Colorear estado
        cell_estado = ws.cell(row=fila, column=12)
        if p.estado == "Mora":
            cell_estado.font = Font(color=COLOR_ROJO, bold=True, size=9, name="Calibri")
        elif p.estado == "Atrasado":
            cell_estado.font = Font(color="D68910", bold=True, size=9, name="Calibri")

        if vencidas > 0:
            ws.cell(row=fila, column=10).font = Font(color=COLOR_ROJO, bold=True, size=9, name="Calibri")

        total_capital += p.capital
        total_saldo += saldo

    # Totales
    fila_tot = 5 + len(prestamos) + 1
    ws.cell(row=fila_tot, column=3, value="TOTALES:").font = Font(bold=True, size=10)
    ws.cell(row=fila_tot, column=4, value=total_capital).number_format = '#,##0'
    ws.cell(row=fila_tot, column=7, value=total_saldo).number_format = '#,##0'
    ws.cell(row=fila_tot, column=4).font = Font(bold=True, color=COLOR_VERDE, size=10)
    ws.cell(row=fila_tot, column=7).font = Font(bold=True, color=COLOR_ROJO, size=10)
>>>>>>> 7761f488b2aa6200974f069ea5072699c6dbd1e5

    anchos = [28, 14, 14, 14, 14, 14, 14, 10, 8, 10, 16, 12]
    for i, a in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = a

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def reporte_resumen_zonas(db: Session, empresa_id: int = None, fecha_desde: datetime.date = None, fecha_hasta: datetime.date = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen por Zona"
    ws.sheet_view.showGridLines = False

<<<<<<< HEAD
    empresa = "CreditosPro"
    try:
        from app.database import ConfiguracionApp
        _cfg2 = db.query(ConfiguracionApp).filter(ConfiguracionApp.empresa_id==empresa_id).first() if empresa_id else None
        if _cfg2 and _cfg2.empresa_nombre: empresa = _cfg2.empresa_nombre
    except Exception:
        empresa = "CreditosPro"
=======
    from app.database import ConfiguracionApp
    config = db.query(ConfiguracionApp).first()
    empresa = config.empresa_nombre if config else "CreditosPro"
>>>>>>> 7761f488b2aa6200974f069ea5072699c6dbd1e5

    encabezado_reporte(
        ws, "RESUMEN POR ZONA",
        f"Período: {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')}",
        empresa
    )
    ws.merge_cells("A1:I1")
    ws.merge_cells("A2:I2")
    ws.merge_cells("A3:I3")

    zq = db.query(Zona).filter(Zona.activa == True)
    if empresa_id:
        zq = zq.filter(Zona.empresa_id == empresa_id)
    zonas = zq.all()
<<<<<<< HEAD
    zona_ids = [z.id for z in zonas]

    labels = ["Zona", "Cobrador", "Clientes", "Préstamos", "Capital Total", "Total Cobrado", "Saldo Pendiente", "Cuotas Vencidas", "Estado"]
    estilo_header(ws, 5, list(range(1, 10)), labels)

    # ── Precálculos bulk para evitar N+1 ────────────────────────────────────
    # Clientes activos por zona
    clientes_por_zona = {}
    if zona_ids:
        for row in db.query(Cliente.zona_id, func.count(Cliente.id).label('cnt')) \
                .filter(Cliente.zona_id.in_(zona_ids), Cliente.activo == True) \
                .group_by(Cliente.zona_id).all():
            clientes_por_zona[row.zona_id] = row.cnt

    # Préstamos activos por zona
    prestamos_por_zona = {}
    if zona_ids:
        for row in db.query(Prestamo.zona_id, func.count(Prestamo.id).label('cnt')) \
                .filter(Prestamo.zona_id.in_(zona_ids), Prestamo.estado == "Activo") \
                .group_by(Prestamo.zona_id).all():
            prestamos_por_zona[row.zona_id] = row.cnt

    # Capital total por zona
    capital_por_zona = {}
    if zona_ids:
        for row in db.query(Prestamo.zona_id, func.sum(Prestamo.capital).label('total')) \
                .filter(Prestamo.zona_id.in_(zona_ids)) \
                .group_by(Prestamo.zona_id).all():
            capital_por_zona[row.zona_id] = row.total

    # Total cobrado por zona en el período
    cobrado_por_zona = {}
    if zona_ids:
        for row in db.query(Cobro.zona_id, func.sum(Cobro.valor_cobrado).label('total')) \
                .filter(Cobro.zona_id.in_(zona_ids),
                        Cobro.fecha >= fecha_desde,
                        Cobro.fecha <= fecha_hasta) \
                .group_by(Cobro.zona_id).all():
            cobrado_por_zona[row.zona_id] = row.total

    # Cuotas vencidas por zona
    vencidas_por_zona = {}
    if zona_ids:
        for row in db.query(Prestamo.zona_id, func.count(Cuota.id).label('cnt')) \
                .join(Cuota, Cuota.prestamo_id == Prestamo.id) \
                .filter(Prestamo.zona_id.in_(zona_ids), Cuota.estado == "Vencida") \
                .group_by(Prestamo.zona_id).all():
            vencidas_por_zona[row.zona_id] = row.cnt

    for idx, zona in enumerate(zonas):
        capital_total = capital_por_zona.get(zona.id, 0)
        cobrado = cobrado_por_zona.get(zona.id, 0)
        saldo = capital_total - cobrado

        fila = 6 + idx
        vals = [
            zona.nombre,
            zona.cobrador_nombre or "—",
            clientes_por_zona.get(zona.id, 0),
            prestamos_por_zona.get(zona.id, 0),
            capital_total,
            cobrado,
            saldo,
            vencidas_por_zona.get(zona.id, 0),
=======
    labels = ["Zona", "Cobrador", "Clientes", "Préstamos", "Capital Total", "Total Cobrado", "Saldo Pendiente", "Cuotas Vencidas", "Estado"]
    estilo_header(ws, 5, list(range(1, 10)), labels)

    for idx, zona in enumerate(zonas, start=1):
        clientes_cnt = db.query(Cliente).filter(Cliente.zona_id == zona.id, Cliente.activo == True).count()
        prestamos_cnt = db.query(Prestamo).filter(Prestamo.zona_id == zona.id, Prestamo.estado == "Activo").count()
        capital = db.query(Prestamo).filter(Prestamo.zona_id == zona.id).all()
        capital_total = sum(p.capital for p in capital)

        cobros = db.query(Cobro).filter(
            Cobro.zona_id == zona.id,
            Cobro.fecha >= fecha_desde,
            Cobro.fecha <= fecha_hasta,
        ).all()
        cobrado = sum(c.valor_cobrado for c in cobros)

        cuotas_venc = db.query(Cuota).join(Prestamo).filter(
            Prestamo.zona_id == zona.id,
            Cuota.estado == "Vencida"
        ).count()

        saldo = capital_total - cobrado

        fila = 5 + idx
        vals = [
            zona.nombre,
            zona.cobrador_nombre or "—",
            clientes_cnt,
            prestamos_cnt,
            capital_total,
            cobrado,
            saldo,
            cuotas_venc,
>>>>>>> 7761f488b2aa6200974f069ea5072699c6dbd1e5
            "Activa" if zona.activa else "Inactiva",
        ]
        estilo_fila(ws, fila, vals, par=(idx % 2 == 0))
        for col in [5, 6, 7]:
            ws.cell(row=fila, column=col).number_format = '#,##0'

    anchos = [16, 24, 10, 12, 18, 18, 18, 14, 10]
    for i, a in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = a

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
